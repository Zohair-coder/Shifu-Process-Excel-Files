from openpyxl import Workbook, load_workbook
import os
import re
import csv

ROOT_DIR = "Archive"
CURRENT_ROW_NUMBER = 1


def main():
    workbook = initialize_worksheet()
    process_all_files(workbook)
    workbook.save(filename="output.xlsx")


def process_all_files(workbook):
    directories = os.listdir(ROOT_DIR)
    directories.sort()
    for directory in directories:
        if is_unnecessary_directory(directory):
            continue
        process_ff_directory(workbook, directory)


def process_ff_directory(workbook, directory):
    add_title_to_sheet(workbook, directory)
    ff_dir_path = os.path.join(ROOT_DIR, directory)
    ff_dir_contents = os.listdir(ff_dir_path)
    for content in ff_dir_contents:
        final_dir_path = os.path.join(ff_dir_path, content)
        if os.path.isdir(final_dir_path):
            final_dir_contents = os.listdir(final_dir_path)
            for file in final_dir_contents:
                final_file_path = os.path.join(final_dir_path, file)
                convert_csv_to_excel(final_file_path)
            final_dir_contents = os.listdir(final_dir_path)
            final_dir_contents.sort(key=lambda x: int(get_specimen_number(x)))
            for file in final_dir_contents:
                final_file_path = os.path.join(final_dir_path, file)
                process_file(workbook, final_file_path)
    add_blank_row(workbook)


def add_blank_row(workbook):
    sheet = workbook.active
    global CURRENT_ROW_NUMBER
    sheet.cell(row=CURRENT_ROW_NUMBER, column=1).value = ""
    CURRENT_ROW_NUMBER += 1


def add_title_to_sheet(workbook: Workbook, directory):
    sheet = workbook.active
    global CURRENT_ROW_NUMBER
    sheet.cell(row=CURRENT_ROW_NUMBER, column=1).value = directory
    CURRENT_ROW_NUMBER += 1


def process_file(workbook, final_file_path):
    if final_file_path.endswith(".xlsx"):
        specimen_number = get_specimen_number(final_file_path)
        value_1 = get_value_1(final_file_path)
        value_2 = get_value_2(final_file_path)
        value_3 = get_value_3(final_file_path)

        add_values_to_sheet(
            workbook, specimen_number, value_1, value_2, value_3)


def convert_csv_to_excel(final_file_path):
    if final_file_path.endswith(".csv"):
        convert_to_excel(final_file_path)
        os.remove(final_file_path)


def get_specimen_number(file):
    if "/" in file:
        file = file.split("/")[-1]
    matches = re.search(
        r'Specimen_RawData_(\d+).xlsx', file)
    specimen_number = matches.group(1)
    return specimen_number


def is_unnecessary_directory(directory):
    return directory == ".DS_Store"


def convert_to_excel(path):
    wb = Workbook()
    ws = wb.active
    with open(path) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)
    wb.save(filename=path.replace(".csv", ".xlsx"))


def initialize_worksheet():
    workbook = Workbook()
    return workbook


def get_data_row(sheet):
    for row in sheet.iter_rows():
        if row[0].value == "Time":
            return row[0].row + 2


def get_value_1(path):
    print(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    curr_max = float("-inf")
    data_row = get_data_row(sheet)
    for row in sheet.iter_rows(min_row=data_row):
        value = float(row[2].value)
        curr_max = max(curr_max, value)
    return curr_max


def get_value_2(path):
    value_1 = get_value_1(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    data_row = get_data_row(sheet)
    for row in sheet.iter_rows(min_row=data_row):
        value = float(row[2].value)
        if value == value_1:
            return row[1].row


def get_value_3(path):
    value_2 = get_value_2(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    data_row = get_data_row(sheet)
    for row in sheet.iter_rows(min_row=data_row):
        value = float(row[2].value)
        if value >= 0.1:
            value_to_subtract = float(row[1].value)
            print(sheet.cell(row=value_2, column=2).value)
            print(value_to_subtract)
            print(float(sheet.cell(row=value_2, column=2).value) - value_to_subtract)
            return float(sheet.cell(row=value_2, column=2).value) - value_to_subtract


def add_values_to_sheet(workbook: Workbook, specimen_number, value_1, value_2, value_3):
    sheet = workbook.active
    global CURRENT_ROW_NUMBER
    sheet.cell(row=CURRENT_ROW_NUMBER, column=1).value = "S" + specimen_number
    sheet.cell(row=CURRENT_ROW_NUMBER, column=2).value = value_1
    sheet.cell(row=CURRENT_ROW_NUMBER, column=3).value = value_2
    sheet.cell(row=CURRENT_ROW_NUMBER, column=4).value = value_3
    CURRENT_ROW_NUMBER += 1


if __name__ == "__main__":
    main()
